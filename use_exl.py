import openpyxl

"""
book = openpyxl.load_workbook("stat_402301.xlsx")
# 첫 번째 방법
# print(book.sheetnames)
# print(book["Sheet0"])

# 두 번째 방법
sheet = book.worksheets[0]
for row in sheet.rows:
    for data in row:
        print(data.value, end="")
    print("",end="\n")
"""
workbook = openpyxl.Workbook()
sheet = workbook.active

# 데이터 쓰기
sheet["A1"] = "테스트 파일"
sheet["A2"] = "안녕 하세요"
sheet.merge_cells("A1:C1")
sheet["A1"].font = openpyxl.styles.Font(size=20,color="FF0000")
workbook.save("newFile.xlsx")